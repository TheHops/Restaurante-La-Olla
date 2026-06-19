# Se importa la funcion para las respuestas del sitio web
from django.db.models import Case, When, Count, Sum
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.views.decorators.http import require_POST
from Application.models  import AreaMesa, Arqueo, Cargo, Mesa, Orden, Platillo, TipoPlatillo, Usuario, DetalleOrden, MesasPorOrden, OTP
from django.contrib import messages
from django.http import JsonResponse
from datetime import datetime, timedelta, time
from django.db.models import Prefetch

from django.utils import timezone

import traceback

from datetime import timezone as py_timezone
from django.core.mail import send_mail

import secrets
import string
import re
import threading

from django.conf import settings

from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError

import io
from django.http import HttpResponse
from django.core.management import call_command
from django.apps import apps

# Librerías para los formatos visuales
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

User = get_user_model()

def index(request):
    # Direccion que me llevará al login por defecto
    if request.user.is_authenticated:
        if request.user.IdCargo.Nombre == "Armador":
            return redirect("OrdenesPendientes/")
        elif request.user.IdCargo.Nombre == "Mesero":
            return redirect("venta/")
        
        # Si el usuario ya inició sesión entonces entrará directamente al sistema sin necesidad de volver a iniciar sesión
        return render(request, "inicio.html", {"Cargo": request.user.IdCargo.Nombre, "User": request.user})
    else:
        # Si no lo ha hecho entonces deberá iniciar sesión
        return render(request, "login.html")
    
def DebeCambiarPass(request):
    if not request.user.is_authenticated:
        return redirect("/")
    
    if request.method != "GET":
        return JsonResponse({"status": "error", "message": "Método no permitido"}, status=405)
    
    return JsonResponse({"status": "ok", "DebeCambiarPass": request.user.DebeCambiarPass})

def loginUser(request):

    if request.method == "POST":

        entrada = request.POST.get("txtUsername")
        pass_word = request.POST.get("txtPassword")

        # Primero detectamos si escribieron un correo
        es_correo = "@" in entrada and "." in entrada

        user_obj = None

        if es_correo:
            try:
                user_obj = User.objects.get(email=entrada)
                username_real = user_obj.username
            except User.DoesNotExist:
                messages.error(request, "El correo ingresado no está asociado a ninguna cuenta.")
                return render(request, "login.html", {"is_for_incorrect_login":True, "message": "El correo ingresado no está asociado a ninguna cuenta", "icon": "error"})
        else:
            username_real = entrada

        user = authenticate(request, username=username_real, password=pass_word)

        if user is None:
            messages.error(request, "Credenciales incorrectas.")
            return render(request, "login.html", {"is_for_incorrect_login":True, "message": "Credenciales incorrectas", "icon": "error"})

        if user.EsActivo != "1":
            messages.error(request, "Tu cuenta está inactiva.")
            return render(request, "login.html", {"is_for_incorrect_login":True, "message": "La cuenta está inactiva", "icon": "error"})

        login(request, user)
        return redirect("/")

    return render(request, "login.html", {"is_for_incorrect_login":False, "message": "", "icon": ""})

def logoutUser(request):
    logout(request)
    return redirect("/")

#region GraficasOrdenes

dias_semana_es = {
    0: "Lunes",
    1: "Martes",
    2: "Miércoles",
    3: "Jueves",
    4: "Viernes",
    5: "Sábado",
    6: "Domingo",
}

def GraficarOrdenes(request):
    if not request.user.is_authenticated:
        return render(request, "login.html")
    
    if request.user.IdCargo.Nombre == "Armador" or request.user.IdCargo.Nombre == "Mesero":
        return redirect("/")

    try:
        cargo_usuario = request.user.IdCargo.Nombre
        data = {}

        # 1. Datos para el Administrador (7 días)
        if cargo_usuario == "Administrador":
            hoy_local = timezone.localdate()
            dias = [(hoy_local - timedelta(days=i)) for i in range(6, -1, -1)]
            inicio_rango = timezone.make_aware(datetime.combine(dias[0], time.min))
            fin_rango = timezone.make_aware(datetime.combine(dias[-1], time.max))
            
            ventas_7_dias = Orden.objects.filter(
                Estado="0", EsActivo="1", 
                UltimaModificacion__range=(inicio_rango, fin_rango)
            )
            
            mapeo_ventas = {}
            for factura in ventas_7_dias:
                fecha = factura.UltimaModificacion.astimezone(timezone.get_current_timezone()).date()
                mapeo_ventas[fecha] = mapeo_ventas.get(fecha, 0) + float(factura.TotalPagar)
            
            data["ingresos_v"] = [mapeo_ventas.get(dia, 0) for dia in dias]
            data["labels_x"] = [dias_semana_es[dia.weekday()] for dia in dias]
            data["resumen"] = obtener_metricas_resumen()

        # 2. Datos para el Cajero (Hoy por horas)
        elif cargo_usuario == "Cajero":
            labels_h, valores_h = obtener_ventas_por_horas()
            data["labels_x"] = labels_h
            data["ingresos_v"] = valores_h
            data["cajero_stats"] = obtener_metricas_cajero()

        # 3. Métodos de Pago (Común para ambos, pero podrías filtrar 
        # que el cajero solo vea lo de HOY y el admin 30 días)
        dias_metodos = 1 if cargo_usuario == "Cajero" else 30
        data["metodos_labels"], data["metodos_valores"] = obtener_stats_metodos_pago(dias_metodos)

        return JsonResponse(data)
    except Exception as ex:
        print("\n\n############### E X C E P C I Ó N ###############")
        print(traceback.format_exc())
        print("#####################################################\n\n")
        return JsonResponse({"error": str(ex)}, status=500)
    
def obtener_stats_metodos_pago(dias_atras=30):
    """
    Retorna labels y valores de métodos de pago en un rango de días.
    """
    hoy_local = timezone.localdate()
    fecha_inicio = hoy_local - timedelta(days=dias_atras)
    
    # Límites con zona horaria
    inicio_dt = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
    fin_dt = timezone.make_aware(datetime.combine(hoy_local, datetime.max.time()))
    

    # Consulta agrupada
    stats = (
        Orden.objects.filter(
            Estado="0", 
            EsActivo="1",
            UltimaModificacion__range=(inicio_dt, fin_dt)
        )
        .values('MetodoPago')
        .annotate(total=Count('Id'))
        .order_by('-total')
    )

    # Diccionario de traducción según tus modelos
    nombres_map = {
        "1": "Efectivo",
        "2": "Tarjeta",
        "3": "Transferencia",
        "4": "Efectivo y tarjeta",
        "5": "N/A"
    }

    labels = []
    valores = []

    for item in stats:
        nombre = nombres_map.get(item['MetodoPago'], "Otro")
        labels.append(nombre)
        valores.append(item['total'])

    return labels, valores

def obtener_ventas_por_horas():
    hoy = timezone.localdate()
    inicio_hoy = timezone.make_aware(datetime.combine(hoy, time.min))
    fin_hoy = timezone.make_aware(datetime.combine(hoy, time.max))

    bloques = [6, 8, 10, 12, 14, 16, 18, 20, 22]
    labels_horas = ["6AM", "8AM", "10AM", "12MD", "2PM", "4PM", "6PM", "8PM", "10PM"]
    
    # Inicializamos un diccionario para acumular: {6: 0, 8: 0, ...}
    acumulador = {hora: 0.0 for hora in bloques}

    ordenes_hoy = Orden.objects.filter(
        Estado="0", 
        EsActivo="1", 
        UltimaModificacion__range=(inicio_hoy, fin_hoy)
    )

    tz_local = timezone.get_current_timezone()

    for orden in ordenes_hoy:
        # 1. Convertimos la hora de la orden a la zona horaria de Nicaragua
        hora_local = orden.UltimaModificacion.astimezone(tz_local).hour
        
        # 2. Buscamos en qué bloque cae (si es 7:45, entra en el bloque del 6 porque 6 <= 7 < 8)
        # O en tu caso, si usas bloques de 2 horas:
        for hora_bloque in bloques:
            if hora_bloque <= hora_local < (hora_bloque + 2):
                acumulador[hora_bloque] += float(orden.TotalPagar)
                break

    # Convertimos el diccionario a la lista que espera la gráfica
    ventas_por_hora = [acumulador[h] for h in bloques]
    
    return labels_horas, ventas_por_hora

def obtener_metricas_resumen():
    hoy = timezone.localdate()
    hace_30_dias = hoy - timedelta(days=30)
    
    # Límites para los filtros
    inicio_hoy = timezone.make_aware(datetime.combine(hoy, datetime.min.time()))
    fin_hoy = timezone.make_aware(datetime.combine(hoy, datetime.max.time()))
    inicio_30 = timezone.make_aware(datetime.combine(hace_30_dias, datetime.min.time()))
    # inicio_hoy = datetime.combine(hoy, datetime.min.time())
    # fin_hoy = datetime.combine(hoy, datetime.max.time())
    # inicio_30 = datetime.combine(hace_30_dias, datetime.min.time())

    # 1. Total del día (Ventas + Propinas)
    stats_hoy = Orden.objects.filter(
        Estado="0", EsActivo="1",
        UltimaModificacion__range=(inicio_hoy, fin_hoy)
    ).aggregate(
        total_ventas=Sum('Total') + Sum('Descuento'),
        total_propinas=Sum('Propina')
    )

    # 2. Total últimos 30 días
    stats_30 = Orden.objects.filter(
        Estado="0", EsActivo="1",
        UltimaModificacion__range=(inicio_30, fin_hoy)
    ).aggregate(
        total_periodo=Sum('Total') + Sum('Descuento'),
        total_propinas_periodo=Sum('Propina'),
        total_gran_total=Sum('TotalPagar'),
        total_cantidad_ordenes=Count('Id')
    )

    return {
        "hoy_total": float(stats_hoy['total_ventas'] or 0),
        "hoy_propinas": float(stats_hoy['total_propinas'] or 0),
        "mes_total": float(stats_30['total_periodo'] or 0),
        "mes_propinas": float(stats_30['total_propinas_periodo'] or 0),
        "mes_gran_total": float(stats_30['total_gran_total'] or 0),
        "mes_cantidad_ordenes": float(stats_30['total_cantidad_ordenes'] or 0),
    }
    
def obtener_metricas_cajero():
    hoy = timezone.localdate()
    inicio_hoy = timezone.make_aware(datetime.combine(hoy, time.min))
    fin_hoy = timezone.make_aware(datetime.combine(hoy, time.max))

    # Filtramos órdenes de hoy
    ordenes_hoy = Orden.objects.filter(Estado="0", EsActivo="1", UltimaModificacion__range=(inicio_hoy, fin_hoy))

    # Métricas específicas (Efectivo es ID "1", Tarjeta es ID "2")
    efectivo = ordenes_hoy.filter(MetodoPago="1").aggregate(v=Sum('TotalPagar'), p=Sum('Propina'))
    tarjeta = ordenes_hoy.filter(MetodoPago="2").aggregate(v=Sum('TotalPagar'), p=Sum('Propina'))

    return {
        "hoy_efectivo_total": float(efectivo['v'] or 0),
        "hoy_efectivo_propina": float(efectivo['p'] or 0),
        "hoy_tarjeta_total": float(tarjeta['v'] or 0),
        "hoy_tarjeta_propina": float(tarjeta['p'] or 0),
    }

#endregion GraficarOrdenes

#region FiltrarOrdenes

OPCIONES_DISPONIBLES = {
    "Administrador": ["0", "1", "2", "3", "4", "5"],
    "Mesero": ["1", "4", "3", "7"],
    "Armador": ["1", "4", "6"],
    "Cajero": ["0", "2", "3"]
}

VALORES_POR_DEFECTO = {
    "Administrador": "5",
    "Mesero": "7",
    "Armador": "6",
    "Cajero": "3"
}

def validar_filtro_por_cargo(cargo, valor_filtro):
    opciones = OPCIONES_DISPONIBLES.get(cargo, [])
    default = VALORES_POR_DEFECTO.get(cargo, None)

    if valor_filtro in opciones:
        return valor_filtro

    return default

def FiltrarOrdenes(request):
    if not request.user.is_authenticated:
        return render(request, "login.html")

    if request.method != "GET":
        return JsonResponse({"status": "error", "message": "Método no permitido"}, status=405)

    try:
        EstadoOrden = request.GET.get("SelectFiltrarOrdenes")
        cargo_usuario = request.user.IdCargo.Nombre

        validar_cargo(request, cargo_usuario)

        # Validación segura del filtro
        EstadoOrden = validar_filtro_por_cargo(cargo_usuario, EstadoOrden)

        # Query base
        ordenes = Orden.objects.select_related(
            'IdUsuario'
        ).prefetch_related(
            Prefetch('Detalles', queryset=DetalleOrden.objects.order_by('-EsActivo')),
            Prefetch('Mesas', queryset=MesasPorOrden.objects.filter(EsActivo="1").select_related('IdMesa')
    )
        ).filter(EsActivo="1")

        # FILTRO ADICIONAL PARA MESERO
        if cargo_usuario == "Mesero":
            ordenes = ordenes.filter(IdUsuario=request.user)

        # FILTROS POR ESTADO
        if EstadoOrden == "5": 
            ordenes = ordenes.order_by(
                Case(
                    When(Estado='1', then=0),
                    When(Estado='4', then=1),
                    When(Estado='3', then=2),
                    When(Estado='0', then=3),
                    When(Estado='2', then=4),
                ),
                '-UltimaModificacion'
            )

        elif EstadoOrden == "6":
            ordenes = ordenes.filter(
                Estado__in=["1", "4"]
            ).order_by(
                Case(
                    When(Estado='1', then=0),
                    When(Estado='4', then=1),
                ),
                '-UltimaModificacion'
            )
            
        elif EstadoOrden == "7":
            ordenes = ordenes.filter(
                Estado__in=["1", "4", "3"]
            ).order_by(
                Case(
                    When(Estado='1', then=0),
                    When(Estado='4', then=1),
                    When(Estado='3', then=2),
                ),
                '-UltimaModificacion'
            )

        else:
            ordenes = ordenes.filter(
                Estado=EstadoOrden
            ).order_by('-UltimaModificacion')

        platillos = Platillo.objects.all().values()

        contexto = {
            "Ordenes": ordenes,
            "Platillos": platillos,
            "CargoUsuario": cargo_usuario
        }

        return render(request, "ordenesFiltradas.html", contexto)

    except Exception as ex:
        print("\n\n############### E X C E P C I Ó N ###############")
        print(traceback.format_exc())
        print("#####################################################\n\n")
        
        return JsonResponse({'error': str(ex)}, status=500)

#endregion FiltrarOrdenes

#region Cargo

def consultar_cargo(request):
    usuario = request.user

    if not usuario.is_authenticated:
        return render(request, "login.html")

    if usuario.IdCargo is None:
        return render(request, "login.html")

    return JsonResponse({"status": "ok", "CargoUsuario": usuario.IdCargo.Nombre})

def validar_cargo(request, cargo_recibido):

    # Verificar autenticación
    if not request.user.is_authenticated:
        return render(request, "login.html")

    cargo_usuario = request.user.IdCargo.Nombre if request.user.IdCargo else ""

    # Comparación estricta
    if cargo_usuario != cargo_recibido:
        return JsonResponse({
            "status": "error",
            "message": "Cargo inválido o manipulado desde el cliente"
        }, status=403)

    # Si coincide → todo bien
    return None

#endregion Cargo

#region Correo

def enviar_correo_hilo(asunto, mensaje, remitente, destinatarios):
    try:
        send_mail(asunto, mensaje, remitente, destinatarios, fail_silently=False)
    except Exception as e:
        print("\n\n############### E X C E P C I Ó N ###############")
        print(traceback.format_exc())
        print("#####################################################\n\n")

def EnviarCorreo(request):
    if not request.user.is_authenticated:
        return render(request, "login.html")
    
    if request.user.IdCargo.Nombre == "Armador" or request.user.IdCargo.Nombre == "Mesero" or request.user.IdCargo.Nombre == "Cajero":
            return redirect("/")
    
    if request.method == "POST":
        try:
            id_personal = request.POST.get("idPersonal")
            titulo = request.POST.get("tituloCorreo")
            mensaje = request.POST.get("mensajeCorreo")
            
            usuario = Usuario.objects.get(Id=id_personal)
            mensaje = f"Hola {usuario.Nombres},\n\n" + f"{mensaje}"

            # enviado = send_mail(
            #     subject=titulo,
            #     message=mensaje,
            #     from_email=settings.DEFAULT_FROM_EMAIL,
            #     recipient_list=[usuario.email],
            #     fail_silently=False,
            # )
            
            thread = threading.Thread(
                target=enviar_correo_hilo,
                args=(titulo, mensaje, settings.DEFAULT_FROM_EMAIL, [usuario.email])
            )
            
            thread.start()
            
            return JsonResponse({'status': 'ok', 'message': '¡El correo ya se encuentra en proceso!'})
            
        except Exception as e:
            return JsonResponse({"ok": False, "message": f"Error al enviar el correo: {str(e)}"})

#endregion Correo

#region ForgotPassword

def ejecutar_envio_correo(asunto, mensaje, destinatario):
    """Esta función corre en un hilo separado."""
    try:
        send_mail(
            subject=asunto,
            message=mensaje,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[destinatario],
            fail_silently=False,
        )
        
        print(f"DEBUG: Correo enviado con éxito a {destinatario}")
    except Exception as e:
        print(f"ERROR al enviar correo en segundo plano: {str(e)}")

def generar_otp(longitud=6):
    caracteres = string.digits
    otp = ''.join(secrets.choice(caracteres) for _ in range(longitud))
    return otp

def generar_crear_otp(usuario):
    # Invalidar OTP anteriores no usados
    OTP.objects.filter(
        Usuario=usuario,
        Usado=False
    ).update(Usado=True)
    
    codigo = generar_otp()
    expiracion = timezone.now() + timedelta(minutes=2)

    OTP.objects.create(
        Usuario=usuario,
        Codigo=codigo,
        FechaExpiracion=expiracion
    )

    return codigo, expiracion

def enviar_otp_correo(usuario, otp):
    try:
        asunto = "Código OTP para restablecer contraseña"
        mensaje = f"Hola {usuario.Nombres},\n\nTu código OTP para restablecer tu contraseña es:\n\n{otp}\n\nEste código expirará en 2 minutos. Si no solicitaste este OTP, ignora este mensaje."
        
        hilo = threading.Thread(
            target=ejecutar_envio_correo,
            args=(asunto, mensaje, usuario.email)
        )
        
        # Lo iniciamos (esto no bloquea la ejecución)
        hilo.start()
        
        # Respondemos de inmediato
        return {
            "ok": True, 
            "message": "Se está procesando el envío del código OTP a tu correo."
        }
    except Exception as e:
        return {"ok": False, "message": f"Error al enviar el OTP: {str(e)}"}

# Método que sirve para renderizar los elementos necesarios para cambiar contraseña desde login
def ForgotPassword(request):
    if request.method == "GET":
        return render(request, "inicio_forgot_password.html")

@require_POST
def ValidateEmailForgotPass(request):

    correo = request.POST.get("txtCorreoForgotPass")

    if not correo:
        return JsonResponse({
            "ok": False,
            "message": "Debes ingresar un correo electrónico"
        }, status=400)

    try:
        usuario = Usuario.objects.select_related("IdCargo").get(email=correo)
    except Usuario.DoesNotExist:
        return JsonResponse({
            "ok": False,
            "message": "No existe ningún usuario asociado a este correo"
        }, status=400)

    try:
        if usuario.EsActivo != "1":
            return JsonResponse({
                "ok": False,
                "message": "El usuario asociado a este correo está inactivo"
            }, status=400)

        if not usuario.IdCargo or usuario.IdCargo.Nombre != "Administrador":
            return JsonResponse({
                "ok": False,
                "message": "Esta función solo está disponible para administradores"
            }, status=400)
        
        otp, expiracion = generar_crear_otp(usuario)
        
        segundos_restantes = int((expiracion - timezone.now()).total_seconds())
        
        resultado_envio = enviar_otp_correo(usuario, otp)

        if not resultado_envio["ok"]:
            return JsonResponse({
                "ok": False,
                "message": resultado_envio["message"]
            }, status=400)
        
        contexto = {
            "Usuario": usuario,
            "segundos_restantes": segundos_restantes
        }

        return render(request, "ingresar_otp_forgot_password.html", contexto)
    except Exception as ex:
        print("\n\n#################### E X C E P C I O N ########################")
        print("----------------------------'ForgotPass'--------------------------")
        print(traceback.format_exc(ex))
        print("###################################################################\n\n")
        
        return JsonResponse({
            "ok": False,
            "message": "Ocurrió un error al generar el OTP"
        }, status=500)
        
@require_POST
def ReenviarOTPForgotPass(request):
    user_id = request.POST.get("idUsuario")

    try:
        usuario = Usuario.objects.get(Id=user_id)

        otp, expiracion = generar_crear_otp(usuario)
        envio = enviar_otp_correo(usuario, otp)

        if not envio["ok"]:
            return JsonResponse({"ok": False, "message": envio["message"]}, status=400)

        segundos = int((expiracion - timezone.now()).total_seconds())

        return JsonResponse({
            "ok": True,
            "segundos": segundos
        })

    except Exception:
        return JsonResponse({
            "ok": False,
            "message": "Error al reenviar OTP"
        }, status=500)
        
@require_POST
def ValidarOTPForgotPass(request):
    otp_ingresado = request.POST.get("otp")
    id_usuario = request.POST.get("id_usuario")

    # Validar que se envíen los datos
    if not otp_ingresado or not id_usuario:
        return JsonResponse({
            "ok": False,
            "message": "Datos incompletos."
        }, status=400)

    # Validar formato del OTP (6 dígitos numéricos)
    if not re.fullmatch(r"\d{6}", otp_ingresado):
        return JsonResponse({
            "ok": False,
            "message": "El OTP ingresado no es válido."
        }, status=400)

    try:
        # Obtener OTP activo del usuario
        otp_obj = OTP.objects.get(
            Usuario__Id=id_usuario,
            Codigo=otp_ingresado,
            Usado=False
        )
    except OTP.DoesNotExist:
        return JsonResponse({
            "ok": False,
            "message": "El OTP es incorrecto o ya fue utilizado."
        }, status=400)

    # Validar expiración
    if timezone.now() > otp_obj.FechaExpiracion:
        return JsonResponse({
            "ok": False,
            "message": "El OTP ha expirado."
        }, status=400)
    
    otp_obj.Usado = True
    otp_obj.save()

    return render(request, "cambiar_pass_forgot_pass.html", {"UserId": id_usuario})

def CambiarPassForgotPass (request):
    if request.method != "POST":
        return JsonResponse({
            "status": "error",
            "message": "Método no permitido"
        }, status=400)
    
    try:
        new_pass = request.POST.get("txtNuevaPassForgotPass")
        verify_pass = request.POST.get("txtVerificarPassForgotPass")
        id_usuario = request.POST.get("userIdValue")

        user = Usuario.objects.get(Id=id_usuario)

        # 🔹 Verificar coincidencia
        if new_pass != verify_pass:
            return JsonResponse({
                "status": "error",
                "message": "Las contraseñas no coinciden"
            }, status=400)

        # 🔹 Validar contraseña con Django
        try:
            validate_password(new_pass, user=request.user)
        except ValidationError as e:
            return JsonResponse({
                "status": "error",
                "message": " ".join(e.messages)
            }, status=400)

        # 🔹 Cambiar contraseña
        user.set_password(new_pass)
        user.DebeCambiarPass = False
        user.save()

        # 🔹 Mantener sesión activa
        update_session_auth_hash(request, user)

        return render(request, "login.html", {"is_for_change_pass":True, "message": "¡La contraseña fue cambiada con éxito!", "icon": "success"})
    except Exception as ex:
        print("\n\n############### E X C E P C I Ó N ###############")
        print(traceback.format_exc())
        print("#####################################################\n\n")
        
        return JsonResponse({
            "status": "error",
            "message": "Ocurrió un error al intentar cambiar la contraseña"
        })

#endregion ForgotPassword

#region Backup

def Respaldo(request):
    # 1. Validaciones de Seguridad y Roles
    if not request.user.is_authenticated:
        return JsonResponse({"status": "error", "message": "Inicie sesión para realizar esta operación."}, status=401)
    
    cargo_usuario = request.user.IdCargo.Nombre if request.user.IdCargo else ""
    if cargo_usuario in ["Armador", "Mesero", "Cajero"]:
        return JsonResponse({"status": "error", "message": "No tiene permisos suficientes para exportar respaldos del sistema."}, status=403)
    
    if request.method == "GET":
        tipo = request.GET.get("Tipo")
        
        # Helper interno para convertir DateTime de UTC a hora local de Nicaragua (UTC-6)
        def ajustar_fecha_local(dt):
            if not dt:
                return "N/A"
            if isinstance(dt, datetime):
                # Si viene con zona horaria (aware), lo pasamos a naive UTC primero
                if timezone.is_aware(dt):
                    dt = timezone.make_naive(dt, py_timezone.utc)
                # Restamos 6 horas para UTC-6
                local_dt = dt - timedelta(hours=6)
                return local_dt.strftime('%d/%m/%Y %I:%M %p')
            return str(dt)

        # Helper para formatear objetos TimeField simples
        def formatear_hora(t):
            if not t:
                return "N/A"
            return t.strftime('%I:%M %p')

        # Obtener fecha actual para el nombre del archivo en UTC-6
        fecha_local_ahora = timezone.localtime(timezone.now())
        fecha_str = fecha_local_ahora.strftime("%Y_%m_%d_%H%M%S")
        
        try:
            # =========================================================================
            # OPCIÓN 1: EXCEL ADMINISTRATIVO - Libro Completo (11 Pestañas)
            # =========================================================================
            if tipo == "1":
                wb = Workbook()
                fuente_cabecera = Font(name='Segoe UI', size=11, bold=True, color="FFFFFF")
                relleno_cabecera = PatternFill(start_color="8E0000", end_color="8E0000", fill_type="solid")
                borde_sutil = Border(bottom=Side(style='thin', color="E5E7EB"), top=Side(style='thin', color="E5E7EB"),
                                     left=Side(style='thin', color="E5E7EB"), right=Side(style='thin', color="E5E7EB"))
                
                # --- PESTAÑA 1: CARGOS ---
                ws1 = wb.active
                ws1.title = "Cargos"
                ws1.append(["ID Cargo", "Nombre del Cargo", "Estado"])
                for c in Cargo.objects.all():
                    ws1.append([c.Id, c.Nombre, "✅ Activo" if c.EsActivo == "1" or getattr(c, 'EsActivo', '1') == "1" else "⛔ Inactivo"])

                # --- PESTAÑA 2: USUARIOS ---
                ws2 = wb.create_sheet(title="Usuarios")
                ws2.append(["ID", "Nombre Completo", "Nombre Usuario", "Email", "Cargo", "Teléfono", "Dirección", "Debe Cambiar Pass", "Estado"])
                for u in Usuario.objects.all():
                    cargo_nom = u.IdCargo.Nombre if u.IdCargo else "Sin Cargo"
                    est = "✅ Activo" if u.EsActivo == "1" else "⛔ Inactivo"
                    ws2.append([u.Id, f"{u.Nombres} {u.Apellidos}", u.username, u.email or 'N/A', cargo_nom, u.Telefono or 'N/A', u.Direccion or 'N/A', "Sí" if u.DebeCambiarPass else "No", est])

                # --- PESTAÑA 3: OTP ---
                ws3 = wb.create_sheet(title="Codigos OTP")
                ws3.append(["ID OTP", "Usuario", "Código", "Fecha Creación", "Fecha Expiración", "Usado"])
                for o in OTP.objects.all():
                    ws3.append([o.Id, o.Usuario.username, o.Codigo, ajustar_fecha_local(o.FechaCreacion), ajustar_fecha_local(o.FechaExpiracion), "Sí" if o.Usado else "No"])

                # --- PESTAÑA 4: AREAS DE MESAS ---
                ws4 = wb.create_sheet(title="Áreas de Mesas")
                ws4.append(["ID Área", "Nombre de Área", "Estado"])
                for am in AreaMesa.objects.all():
                    ws4.append([am.Id, am.Nombre, "✅ Activo" if am.EsActivo == "1" else "⛔ Inactivo"])

                # --- PESTAÑA 5: MESAS ---
                ws5 = wb.create_sheet(title="Mesas")
                ws5.append(["ID Mesa", "Área Asignada", "Número Mesa", "Capacidad", "Estado Ocupación", "Estado Registro"])
                for m in Mesa.objects.all():
                    ws5.append([m.Id, m.IdAreaMesa.Nombre, m.Numero, m.Capacidad or 0, "Disponible" if m.Estado == "1" else "Ocupado", "✅ Activo" if m.EsActivo == "1" else "⛔ Inactivo"])

                # --- PESTAÑA 6: ÓRDENES ---
                ws6 = wb.create_sheet(title="Órdenes")
                ws6.append(["ID Orden", "Atendido Por", "Área/Mesa", "Descripción", "Motivo Anulación", "No. Referencia", "Total", "Total Pagar", "Monto Recibido", "Cambio", "Propina", "Descuento", "Fecha ", "Última Modificación ", "Fue Editada", "Método Pago", "Estado Orden", "Banco"])
                for ord in Orden.objects.all():
                    ws6.append([
                        ord.Id, ord.IdUsuario.username if ord.IdUsuario else "Sistema", ord.AreaDeMesa or 'N/A', ord.Descripcion or '', ord.Motivo or '', ord.NumRef or '',
                        float(ord.Total or 0), float(ord.TotalPagar or 0), float(ord.Monto or 0), float(ord.Cambio or 0), float(ord.Propina or 0), float(ord.Descuento or 0),
                        ajustar_fecha_local(ord.Fecha), ajustar_fecha_local(ord.UltimaModificacion), "Sí" if ord.FueEditada else "No",
                        ord.get_MetodoPago_display(), ord.get_Estado_display(), ord.get_Banco_display() if ord.Banco else 'N/A'
                    ])
                    # Formatos de moneda Córdoba a columnas financieras (G hasta L)
                    for col_idx in range(7, 13):
                        ws6.cell(row=ws6.max_row, column=col_idx).number_format = '"C$"#,##0.00'

                # --- PESTAÑA 7: ARQUEOS DE CAJA ---
                ws7 = wb.create_sheet(title="Arqueos de Caja")
                ws7.append(["ID Arqueo", "Fecha", "Usuario Apertura", "Usuario Cierre", "Hora Apertura", "Hora Cierre", "Monto Inicial", "Monto Teórico", "Monto Real", "Diferencia", "Estado"])
                for ar in Arqueo.objects.all():
                    f_fmt = ar.Fecha.strftime('%d/%m/%Y') if ar.Fecha else 'N/A'
                    u_ap = ar.IdUsuarioApertura.username if ar.IdUsuarioApertura else 'N/A'
                    u_ci = ar.IdUsuarioCierre.username if ar.IdUsuarioCierre else 'N/A'
                    ws7.append([
                        ar.Id, f_fmt, u_ap, u_ci, formatear_hora(ar.HoraApertura), formatear_hora(ar.HoraCierre),
                        float(ar.MontoInicial or 0), float(ar.MontoFinalTeorico or 0), float(ar.MontoFinalReal or 0), float(ar.Diferencia or 0), ar.get_Estado_display()
                    ])
                    ws7.cell(row=ws7.max_row, column=7).number_format = '"C$"#,##0.00'
                    ws7.cell(row=ws7.max_row, column=8).number_format = '"C$"#,##0.00'
                    ws7.cell(row=ws7.max_row, column=9).number_format = '"C$"#,##0.00'
                    ws7.cell(row=ws7.max_row, column=10).number_format = '"C$"#,##0.00'

                # --- PESTAÑA 8: MESAS POR ORDEN ---
                ws8 = wb.create_sheet(title="Mesas Por Orden")
                ws8.append(["ID Relación", "ID Órden", "Número Mesa", "Estado Registro"])
                for mo in MesasPorOrden.objects.all():
                    ws8.append([mo.Id, mo.IdOrden.Id, mo.IdMesa.Numero, "✅ Activo" if mo.EsActivo == "1" else "⛔ Inactivo"])

                # --- PESTAÑA 9: TIPO DE PLATILLOS ---
                ws9 = wb.create_sheet(title="Tipos de Platillos")
                ws9.append(["ID Tipo", "Nombre Categoría", "Estado"])
                for tp in TipoPlatillo.objects.all():
                    ws9.append([tp.Id, tp.Nombre, "✅ Activo" if tp.EsActivo == "1" else "⛔ Inactivo"])

                # --- PESTAÑA 10: PLATILLOS ---
                ws10 = wb.create_sheet(title="Platillos")
                ws10.append(["ID Platillo", "Categoría", "Nombre Platillo", "Precio", "Descripción", "Ruta Imagen", "Estado"])
                for p in Platillo.objects.all():
                    ws10.append([p.Id, p.IdTipoPlatillo.Nombre if p.IdTipoPlatillo else 'General', p.Nombre, float(p.Precio), p.Descripcion or '', str(p.ImagenUrl), "✅ Activo" if p.EsActivo == "1" else "⛔ Inactivo"])
                    ws10.cell(row=ws10.max_row, column=4).number_format = '"C$"#,##0.00'

                # --- PESTAÑA 11: DETALLES DE ÓRDENES ---
                ws11 = wb.create_sheet(title="Detalles de Órdenes")
                ws11.append(["ID Detalle", "ID Orden", "Platillo", "Cantidad", "Precio Venta", "Subtotal Calculated"])
                for do in DetalleOrden.objects.all():
                    subt = float(do.Cantidad or 0) * float(do.PrecioVenta or 0)
                    ws11.append([do.Id, do.IdOrden.Id, do.IdPlatillo.Nombre, do.Cantidad, float(do.PrecioVenta), subt])
                    ws11.cell(row=ws11.max_row, column=5).number_format = '"C$"#,##0.00'
                    ws11.cell(row=ws11.max_row, column=6).number_format = '"C$"#,##0.00'

                # Procesamiento y estilización unificada de hojas
                for sheet in wb.worksheets:
                    sheet.views.sheetView[0].showGridLines = True
                    for col_num in range(1, sheet.max_column + 1):
                        celda = sheet.cell(row=1, column=col_num)
                        celda.font = fuente_cabecera
                        celda.fill = relleno_cabecera
                        celda.alignment = Alignment(horizontal="center", vertical="center")
                    
                    for fila in range(2, sheet.max_row + 1):
                        for col in range(1, sheet.max_column + 1):
                            sheet.cell(row=fila, column=col).border = borde_sutil
                    
                    # Autoajuste de anchos de columna dinámico
                    for col in sheet.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        col_letter = get_column_letter(col[0].column)
                        sheet.column_dimensions[col_letter].width = max(max_len + 4, 13)

                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="Backup_Administrativo_{fecha_str}.xlsx"'
                wb.save(response)
                return response

            # =========================================================================
            # OPCIÓN 2: PDF EJECUTIVO DETALLADO - Reporte Operativo Completo (11 Tablas)
            # =========================================================================
            elif tipo == "2":
                buffer = io.BytesIO()
                # Márgenes de 40pt -> Ancho útil exacto de 532pt
                doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
                story = []
                styles = getSampleStyleSheet()
                
                # Estilos Estructurales del PDF
                titulo_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=16, textColor=colors.HexColor("#111827"), spaceAfter=4)
                sub_style = ParagraphStyle('DocSub', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=colors.HexColor("#4B5563"), spaceAfter=15)
                h2_style = ParagraphStyle('DocH2', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=11, textColor=colors.HexColor("#8E0000"), spaceBefore=14, spaceAfter=6)
                
                # Estilos del contenido interno de tablas (Asegura el auto-wrap de filas largas)
                hdr_style = ParagraphStyle('TableHdr', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, textColor=colors.white, alignment=1)
                cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontName='Helvetica', fontSize=7.5, textColor=colors.HexColor("#1F2937"))
                cell_style_center = ParagraphStyle('TableCellCenter', parent=styles['Normal'], fontName='Helvetica', fontSize=7.5, textColor=colors.HexColor("#1F2937"), alignment=1)
                cell_style_right = ParagraphStyle('TableCellRight', parent=styles['Normal'], fontName='Helvetica', fontSize=7.5, textColor=colors.HexColor("#1F2937"), alignment=2)

                # Estilo de tabla unificado
                tabla_base_style = TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#8E0000")),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E5E7EB")),
                    ('TOPPADDING', (0,0), (-1,-1), 4),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                    ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F9FAFB")]),
                ])

                # Encabezado Principal del Documento
                story.append(Paragraph("SISTEMA DE GESTIÓN LA OLLA - RESPALDO DE DATOS", titulo_style))
                story.append(Paragraph(f"Fecha de generación: {fecha_local_ahora.strftime('%d/%m/%Y %I:%M %p')}", sub_style))
                story.append(Spacer(1, 4))
                
                # --- TABLA 1: CARGOS (532pt) ---
                story.append(Paragraph("1. Roles y Cargos del Personal", h2_style))
                data = [[Paragraph("ID Cargo", hdr_style), Paragraph("Nombre del Cargo", hdr_style), Paragraph("Estado", hdr_style)]]
                for c in Cargo.objects.all():
                    data.append([Paragraph(str(c.Id), cell_style_center), Paragraph(c.Nombre, cell_style), Paragraph("✓ Activo" if getattr(c, 'EsActivo', '1') == "1" else "✗ Inactivo", cell_style_center)])
                t1 = Table(data, colWidths=[80, 332, 120]); t1.setStyle(tabla_base_style); story.append(t1)

                # --- TABLA 2: USUARIOS (532pt) ---
                story.append(Paragraph("2. Historial de Usuarios del Sistema", h2_style))
                data = [[Paragraph("ID", hdr_style), Paragraph("Usuario", hdr_style), Paragraph("Nombre Completo", hdr_style), Paragraph("Cargo", hdr_style), Paragraph("Teléfono", hdr_style), Paragraph("Estado", hdr_style)]]
                for u in Usuario.objects.all():
                    data.append([Paragraph(str(u.Id), cell_style_center), Paragraph(u.username, cell_style), Paragraph(f"{u.Nombres} {u.Apellidos}", cell_style), Paragraph(u.IdCargo.Nombre if u.IdCargo else "Sin Cargo", cell_style), Paragraph(u.Telefono or 'N/A', cell_style_center), Paragraph("✓ Activo" if u.EsActivo == "1" else "✗ Inactivo", cell_style_center)])
                t2 = Table(data, colWidths=[40, 90, 150, 90, 80, 82]); t2.setStyle(tabla_base_style); story.append(t2)

                # --- TABLA 3: OTP (532pt) ---
                story.append(Paragraph("3. Registro de Tokens de Seguridad (OTP)", h2_style))
                data = [[Paragraph("ID OTP", hdr_style), Paragraph("Usuario", hdr_style), Paragraph("Código", hdr_style), Paragraph("Creación ", hdr_style), Paragraph("Expiración ", hdr_style), Paragraph("Usado", hdr_style)]]
                for o in OTP.objects.all():
                    data.append([Paragraph(str(o.Id), cell_style_center), Paragraph(o.Usuario.username, cell_style), Paragraph(o.Codigo, cell_style_center), Paragraph(ajustar_fecha_local(o.FechaCreacion), cell_style_center), Paragraph(ajustar_fecha_local(o.FechaExpiracion), cell_style_center), Paragraph("Sí" if o.Usado else "No", cell_style_center)])
                t3 = Table(data, colWidths=[45, 107, 60, 130, 130, 60]); t3.setStyle(tabla_base_style); story.append(t3)

                # --- TABLA 4: AREAS DE MESAS (532pt) ---
                story.append(Paragraph("4. Distribución de Áreas del Local", h2_style))
                data = [[Paragraph("ID Área", hdr_style), Paragraph("Nombre de Área", hdr_style), Paragraph("Estado", hdr_style)]]
                for am in AreaMesa.objects.all():
                    data.append([Paragraph(str(am.Id), cell_style_center), Paragraph(am.Nombre, cell_style), Paragraph("✓ Activo" if am.EsActivo == "1" else "✗ Inactivo", cell_style_center)])
                t4 = Table(data, colWidths=[80, 332, 120]); t4.setStyle(tabla_base_style); story.append(t4)

                # --- TABLA 5: MESAS (532pt) ---
                story.append(Paragraph("5. Inventario de Mesas", h2_style))
                data = [[Paragraph("ID Mesa", hdr_style), Paragraph("Área local", hdr_style), Paragraph("Número", hdr_style), Paragraph("Capacidad", hdr_style), Paragraph("Ocupación", hdr_style), Paragraph("Estado", hdr_style)]]
                for m in Mesa.objects.all():
                    data.append([Paragraph(str(m.Id), cell_style_center), Paragraph(m.IdAreaMesa.Nombre, cell_style), Paragraph(str(m.Numero), cell_style_center), Paragraph(str(m.Capacidad or 0), cell_style_center), Paragraph("Disponible" if m.Estado == "1" else "Ocupado", cell_style_center), Paragraph("✓ Activo" if m.EsActivo == "1" else "✗ Inactivo", cell_style_center)])
                t5 = Table(data, colWidths=[50, 150, 60, 70, 100, 102]); t5.setStyle(tabla_base_style); story.append(t5)

                # --- TABLA 6: ÓRDENES (Compactado para Vista de Control, 532pt) ---
                story.append(Paragraph("6. Historial General de Órdenes", h2_style))
                data = [[Paragraph("ID", hdr_style), Paragraph("Fecha ", hdr_style), Paragraph("Atendido Por", hdr_style), Paragraph("Mesa/Área", hdr_style), Paragraph("Pago", hdr_style), Paragraph("Total", hdr_style), Paragraph("Estado", hdr_style)]]
                for ord in Orden.objects.all():
                    data.append([Paragraph(str(ord.Id), cell_style_center), Paragraph(ajustar_fecha_local(ord.Fecha), cell_style_center), Paragraph(ord.IdUsuario.username if ord.IdUsuario else "Sistema", cell_style), Paragraph(ord.AreaDeMesa or 'N/A', cell_style), Paragraph(ord.get_MetodoPago_display(), cell_style_center), Paragraph(f"C$ {ord.TotalPagar:,.2f}", cell_style_right), Paragraph(ord.get_Estado_display(), cell_style_center)])
                t6 = Table(data, colWidths=[35, 110, 85, 75, 75, 75, 77]); t6.setStyle(tabla_base_style); story.append(t6)

                # --- TABLA 7: ARQUEOS DE CAJA (532pt) ---
                story.append(Paragraph("7. Libro de Arqueos de Caja", h2_style))
                data = [[Paragraph("ID", hdr_style), Paragraph("Fecha", hdr_style), Paragraph("Apertura/Cierre", hdr_style), Paragraph("M. Inicial", hdr_style), Paragraph("M. Real", hdr_style), Paragraph("Diferencia", hdr_style), Paragraph("Estado", hdr_style)]]
                for ar in Arqueo.objects.all():
                    f_fmt = ar.Fecha.strftime('%d/%m/%Y') if ar.Fecha else 'N/A'
                    u_combo = f"{ar.IdUsuarioApertura.username if ar.IdUsuarioApertura else 'N/A'} / {ar.IdUsuarioCierre.username if ar.IdUsuarioCierre else 'N/A'}"
                    data.append([Paragraph(str(ar.Id), cell_style_center), Paragraph(f_fmt, cell_style_center), Paragraph(u_combo, cell_style), Paragraph(f"C$ {ar.MontoInicial:,.2f}", cell_style_right), Paragraph(f"C$ {ar.MontoFinalReal:,.2f}", cell_style_right), Paragraph(f"C$ {ar.Diferencia or 0:,.2f}", cell_style_right), Paragraph(ar.get_Estado_display(), cell_style_center)])
                t7 = Table(data, colWidths=[35, 65, 115, 80, 80, 80, 77]); t7.setStyle(tabla_base_style); story.append(t7)

                # --- TABLA 8: MESAS POR ORDEN (532pt) ---
                story.append(Paragraph("8. Relación Dinámica de Mesas por Órdenes", h2_style))
                data = [[Paragraph("ID Relación", hdr_style), Paragraph("ID Orden", hdr_style), Paragraph("Número Mesa", hdr_style), Paragraph("Estado Registro", hdr_style)]]
                for mo in MesasPorOrden.objects.all():
                    data.append([Paragraph(str(mo.Id), cell_style_center), Paragraph(str(mo.IdOrden.Id), cell_style_center), Paragraph(str(mo.IdMesa.Numero), cell_style_center), Paragraph("✓ Activo" if mo.EsActivo == "1" else "✗ Inactivo", cell_style_center)])
                t8 = Table(data, colWidths=[100, 120, 140, 172]); t8.setStyle(tabla_base_style); story.append(t8)

                # --- TABLA 9: TIPOS DE PLATILLOS (532pt) ---
                story.append(Paragraph("9. Categorías del Menú (Tipos de Platillo)", h2_style))
                data = [[Paragraph("ID Categoría", hdr_style), Paragraph("Nombre Categoría", hdr_style), Paragraph("Estado", hdr_style)]]
                for tp in TipoPlatillo.objects.all():
                    data.append([Paragraph(str(tp.Id), cell_style_center), Paragraph(tp.Nombre, cell_style), Paragraph("✓ Activo" if tp.EsActivo == "1" else "✗ Inactivo", cell_style_center)])
                t9 = Table(data, colWidths=[90, 300, 142]); t9.setStyle(tabla_base_style); story.append(t9)

                # --- TABLA 10: PLATILLOS (532pt) ---
                story.append(Paragraph("10. Catálogo Maestro de Platillos", h2_style))
                data = [[Paragraph("ID", hdr_style), Paragraph("Categoría", hdr_style), Paragraph("Nombre del Platillo", hdr_style), Paragraph("Precio", hdr_style), Paragraph("Estado", hdr_style)]]
                for p in Platillo.objects.all():
                    data.append([Paragraph(str(p.Id), cell_style_center), Paragraph(p.IdTipoPlatillo.Nombre if p.IdTipoPlatillo else 'General', cell_style), Paragraph(p.Nombre, cell_style), Paragraph(f"C$ {p.Precio:,.2f}", cell_style_right), Paragraph("✓ Activo" if p.EsActivo == "1" else "✗ Inactivo", cell_style_center)])
                t10 = Table(data, colWidths=[40, 120, 190, 90, 92]); t10.setStyle(tabla_base_style); story.append(t10)

                # --- TABLA 11: DETALLES DE ÓRDENES (532pt) ---
                story.append(Paragraph("11. Desglose Analítico de Detalles de Órdenes", h2_style))
                data = [[Paragraph("ID Detalle", hdr_style), Paragraph("ID Orden", hdr_style), Paragraph("Platillo", hdr_style), Paragraph("Cant.", hdr_style), Paragraph("Precio Venta", hdr_style), Paragraph("Subtotal", hdr_style)]]
                for do in DetalleOrden.objects.all():
                    subt = float(do.Cantidad or 0) * float(do.PrecioVenta or 0)
                    data.append([Paragraph(str(do.Id), cell_style_center), Paragraph(str(do.IdOrden.Id), cell_style_center), Paragraph(do.IdPlatillo.Nombre, cell_style), Paragraph(str(do.Cantidad), cell_style_center), Paragraph(f"C$ {do.PrecioVenta:,.2f}", cell_style_right), Paragraph(f"C$ {subt:,.2f}", cell_style_right)])
                t11 = Table(data, colWidths=[55, 55, 202, 40, 90, 90]); t11.setStyle(tabla_base_style); story.append(t11)

                # Construcción estructural del PDF
                doc.build(story)
                buffer.seek(0)
                
                response = HttpResponse(buffer, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="Backup_Operativo_{fecha_str}.pdf"'
                return response

            # =========================================================================
            # OPCIÓN 3: SCRIPT SQL AUTOMÁTICO (Permanece Intacto)
            # =========================================================================
            elif tipo == "3":
                sql_output = f"-- RESPALDO TOTAL - LA OLLA\n-- Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n\n"
                sql_output += "SET FOREIGN_KEY_CHECKS = 0;\n\n"
                modelos_sistema = [Cargo, Usuario, OTP, AreaMesa, Mesa, Orden, Arqueo, MesasPorOrden, TipoPlatillo, Platillo, DetalleOrden]
                for modelo in modelos_sistema:
                    tabla_db = modelo._meta.db_table
                    sql_output += f"-- Datos de la tabla: {tabla_db}\n"
                    registros = modelo.objects.all()
                    for obj in registros:
                        campos, valores = [], []
                        for f in obj._meta.fields:
                            campos.append(f"`{f.column}`")
                            val = getattr(obj, f.attname)
                            if val is None:
                                valores.append("NULL")
                            elif isinstance(val, (int, float)):
                                valores.append(str(val))
                            elif isinstance(val, bool):
                                valores.append("1" if val else "0")
                            else:
                                escapado = str(val).replace("'", "''").replace("\\", "\\\\")
                                valores.append(f"'{escapado}'")
                        sql_output += f"INSERT INTO `{tabla_db}` ({', '.join(campos)}) VALUES ({', '.join(valores)});\n"
                    sql_output += "\n"
                sql_output += "SET FOREIGN_KEY_CHECKS = 1;\n"
                
                response = HttpResponse(sql_output, content_type='text/plain')
                response['Content-Disposition'] = f'attachment; filename="Respaldo_SQL_{fecha_str}.sql"'
                return response

            # =========================================================================
            # OPCIÓN 4: RESPALDO NATIVO JSON (Permanece Intacto)
            # =========================================================================
            elif tipo == "4":
                buffer_json = io.StringIO()
                call_command('dumpdata', 'Application', indent=2, stdout=buffer_json)
                response = HttpResponse(buffer_json.getvalue(), content_type='application/json')
                response['Content-Disposition'] = f'attachment; filename="Respaldo_Estructural_{fecha_str}.json"'
                return response
            
            else:
                return JsonResponse({"status": "error", "message": "Formato de exportación no válido."}, status=400)
                
        except Exception as e:
            return JsonResponse({"status": "error", "message": f"Error interno: {str(e)}"}, status=500)

    return JsonResponse({"status": "error", "message": "Método no permitido."}, status=405)

#endregion Backup

