import os
import traceback
import pdfkit
import json

from datetime import datetime, time
from decimal import Decimal

from jinja2 import Environment, FileSystemLoader

from django.core.exceptions import ValidationError
from django.utils import timezone
from django.utils.formats import date_format
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.db.models import Count, Prefetch, Q, Sum

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

from Application.models import Platillo, TipoPlatillo, Orden, DetalleOrden, AreaMesa, Usuario, MesasPorOrden, Arqueo

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
import io

#region Inicio

def Reportes (request):
    if not request.user.is_authenticated:
        return render(request, "login.html")
    
    if request.user.IdCargo.Nombre == "Armador" or request.user.IdCargo.Nombre == "Mesero":
            return redirect("/")
    
    areas = AreaMesa.objects.filter(EsActivo = "1")
    
    return render(request, "reportes.html", {"Areas": areas, "Cargo": request.user.IdCargo.Nombre, "User": request.user})

#endregion Inicio

#region Ordenes filtradas

def ReportesOrdenesFiltradas (request):
    if not request.user.is_authenticated:
        return render(request, "login.html")
    
    if request.user.IdCargo.Nombre == "Armador" or request.user.IdCargo.Nombre == "Mesero":
        return redirect("/")

    if request.method != "GET":
        return JsonResponse(
            {"status": "error", "message": "Método no permitido"},
            status=405
        )

    try:
        fecha_inicio_str = request.GET.get("FechaInicio")
        fecha_fin_str = request.GET.get("FechaFin")
        estado = request.GET.get("Estado")
        
        areasSeleccionadas = request.GET.get("AreasSeleccionadas", "")
        
        areas_ids = []
        if areasSeleccionadas:
            areas_ids = [int(x) for x in areasSeleccionadas.split(",") if x.isdigit()]

        try:
            ordenes = filtrar_ordenes_fechas_areas(fecha_inicio_str, fecha_fin_str, areas_ids, estado)
        except ValidationError as e:
            return JsonResponse({"status": "error", "message": str(e)})

        contexto = {
            "Ordenes": ordenes,
        }

        return render(
            request,
            "reportes_ordenes_filtradas.html",
            contexto
        )

    except Exception as ex:
        print("\n\n############### E X C E P C I Ó N ###############")
        print(traceback.format_exc())
        print("#####################################################\n\n")
        return JsonResponse(
            {"status": "error", "message": str(ex)},
            status=500
        )

#endregion Ordenes filtradas

#region Mostrar orden

def InicioMostrar(request):
    if not request.user.is_authenticated:
        return render(request, "login.html")
    
    if request.user.IdCargo.Nombre == "Armador" or request.user.IdCargo.Nombre == "Mesero":
        return redirect("/")
    
    idOrden = request.GET.get("IdOrden")
    
    if not idOrden:
        return JsonResponse({"message": "Orden no válida"})

    orden = Orden.objects.prefetch_related(Prefetch('Detalles', queryset=DetalleOrden.objects.filter(EsActivo="1")), Prefetch('Mesas', queryset=MesasPorOrden.objects.filter(EsActivo="1").select_related('IdMesa'))).get(Id=idOrden)
    
    usuario = Usuario.objects.select_related("IdCargo").get(Id=orden.IdUsuario.Id)
    
    contexto = {
        "Orden": orden,
        "Modo": "Mostrar",
        "NombreCreador": usuario.Nombres + " " + usuario.Apellidos,
        "CargoCreador": usuario.IdCargo.Nombre
    }

    return render(request, "detalle_orden_editar.html", contexto)

#endregion Mostrar orden

#region Exportaciones

#region Ordenes

def ExportarOrdenes(request):
    if not request.user.is_authenticated:
        return render(request, "login.html")
    
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Método no permitido"}, status=405)
    
    if request.user.IdCargo.Nombre == "Armador" or request.user.IdCargo.Nombre == "Mesero":
        return redirect("/")
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"status": "error", "message": "JSON inválido"}, status=400)
    
    try:
        fecha_inicio_str = data["FechaInicio"]
        fecha_fin_str = data["FechaFin"]
        areas_ids = data["AreasSeleccionadas"]
        tipo_exportacion = data["TipoExportacion"]
        incluir_detalles = data["IncluirDetalles"]
        estado = data["Estado"]
    except KeyError as e:
        return JsonResponse({"status": "error", "message": f"Falta el campo {str(e)}"}, status=400)
    
    if request.user.IdCargo.Nombre == "Cajero":
        fecha_hoy_local = timezone.localtime(timezone.now()).date()
        
        # Convertimos a string para mantener consistencia con el formato que recibes del JSON
        hoy_str = fecha_hoy_local.isoformat()
        
        fecha_inicio_str = hoy_str
        fecha_fin_str = hoy_str
    
    try:
        ordenes = filtrar_ordenes_fechas_areas(fecha_inicio_str, fecha_fin_str, areas_ids, estado)
    except ValidationError as e:
        return JsonResponse({"status": "error", "message": str(e)})
    
    nombreArchivo = f"ORDENES_{fecha_inicio_str}_{fecha_fin_str}" if not incluir_detalles else f"ORDENES_DETALLES_{fecha_inicio_str}_{fecha_fin_str}"
    
    if tipo_exportacion == "1":
        wb = exportar_excel_ordenes(ordenes, incluir_detalles)
        return descargar_excel(wb, f"{nombreArchivo}.xlsx")
    
    if tipo_exportacion == "2":
        return exportar_pdf_ordenes(ordenes, request, incluir_detalles, nombreArchivo)
    
    return JsonResponse({
            "status": "error",
            "message": "Exportación no implementada"
        }, status=400)

def exportar_excel_ordenes(ordenes, incluir_detalles=False):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Órdenes"

    # --- DEFINICIÓN DE ESTILOS ---
    vino_oscuro = "800000"
    blanco = "FFFFFF"
    gris_totales = "EEEEEE"
    gris_detalles = "F9F9F9"
    negro = "000000"
    
    color_bordes_detalle = "CCCCCC"
    
    # Definir bordes para los detalles
    borde_detalle = Border(
        left=Side(style='thin', color=color_bordes_detalle),
        right=Side(style='thin', color=color_bordes_detalle),
        top=Side(style='thin', color=color_bordes_detalle),
        bottom=Side(style='thin', color=color_bordes_detalle)
    )

    # Estilo Encabezado (Vinotinto)
    if "header_style" not in wb.named_styles:
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, color=blanco)
        header_style.fill = PatternFill("solid", fgColor=vino_oscuro)
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        wb.add_named_style(header_style)

    # Estilo Fila Normal (Blanco, texto normal)
    if "normal_row" not in wb.named_styles:
        normal_row = NamedStyle(name="normal_row")
        normal_row.font = Font(bold=False, color=negro)
        normal_row.fill = PatternFill("solid", fgColor=blanco)
        normal_row.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        wb.add_named_style(normal_row)

    columnas = ["N° Orden", "Estado", "Fecha", "Área/Mesas", "Subtotal", "Propina", "Descuento", "Total pagado", "Método Pago", "Monto", "Cambio", "2do Monto (Pago mixto)", "Creador/Cargo"]
    
    # Título Principal
    ultima_col_letra = get_column_letter(len(columnas))
    ws.merge_cells(f"A1:{ultima_col_letra}1")
    ws["A1"] = "REPORTE DE VENTAS TOTALES"
    ws["A1"].style = "header_style"
    ws.row_dimensions[1].height = 25

    # Encabezados de tabla
    for col_idx, valor in enumerate(columnas, start=1):
        cell = ws.cell(row=2, column=col_idx, value=valor)
        cell.style = "header_style"

    # --- VARIABLES PARA ACUMULAR TOTALES ---
    sum_subtotal = sum_propina = sum_desc = sum_total = sum_monto = sum_cambio = sum_2do = Decimal(0)
    
    # Diccionario para agrupar propinas por usuario
    propinas_por_usuario = {}

    row_idx = 3
    for orden in ordenes:
        mesas = " - ".join(f"#{m.IdMesa.Numero}" for m in orden.Mesas.all())
        area_info = f"{orden.IdAreaDeMesa.Nombre if orden.IdAreaDeMesa else 'N/A'} ({mesas})"
        
        creador = f"{orden.IdUsuario.Nombres} ({orden.IdUsuario.IdCargo.Nombre})"
        
        # --- NUEVA LÓGICA: Acumular propina por usuario ---
        propina_valor = Decimal(str(orden.Propina or 0))
        propinas_por_usuario[creador] = propinas_por_usuario.get(creador, Decimal(0)) + propina_valor
        
        # Datos de la fila
        datos_orden = [
            orden.Id, orden.get_Estado_display(), 
            timezone.localtime(orden.Fecha).replace(tzinfo=None),
            area_info, orden.Total, orden.Propina, orden.Descuento, 
            orden.TotalPagar, orden.get_MetodoPago_display(), 
            orden.Monto, orden.Cambio, orden.SegundoMonto, creador
        ]

        # Llenado de la fila
        for col_idx, valor in enumerate(datos_orden, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.style = "normal_row"
            
            # Formatos numéricos y acumulación (solo si el estado es Pagado/Finalizado o según tu lógica)
            if col_idx in [5, 6, 7, 8, 10, 11, 12]:
                cell.number_format = '"C$"#,##0.00'
                # Acumulamos para el gran total al final
                if valor:
                    if col_idx == 5: sum_subtotal += Decimal(str(valor))
                    if col_idx == 6: sum_propina += Decimal(str(valor))
                    if col_idx == 7: sum_desc += Decimal(str(valor))
                    if col_idx == 8: sum_total += Decimal(str(valor))
                    if col_idx == 10: sum_monto += Decimal(str(valor))
                    if col_idx == 11: sum_cambio += Decimal(str(valor))
                    if col_idx == 12: sum_2do += Decimal(str(valor))
            
            if col_idx == 3:
                cell.number_format = 'dd/mm/yyyy hh:mm AM/PM'

        row_idx += 1

        # --- DETALLES (OPCIONAL) ---
        if incluir_detalles:
            det_headers = ["", "CONSUMIBLE", "CANTIDAD", "PRECIO UNITARIO", "SUBTOTAL"]
            for col_det, text in enumerate(det_headers, start=1):
                if text: # No pintar la primera celda vacía
                    c_h = ws.cell(row=row_idx, column=col_det, value=text)
                    c_h.font = Font(size=8, bold=True, color="555555")
                    c_h.fill = PatternFill("solid", fgColor=gris_detalles)
                    c_h.alignment = Alignment(horizontal="center")
                    c_h.border = borde_detalle
            
            ws.row_dimensions[row_idx].outlineLevel = 1
            row_idx += 1
            
            for det in orden.Detalles.filter(EsActivo="1"):
                # Producto en col B, Cantidad en col C, etc.
                ws.cell(row=row_idx, column=2, value=f" > {det.IdPlatillo.Nombre}").font = Font(size=9, italic=True)
                ws.cell(row=row_idx, column=3, value=det.Cantidad).alignment = Alignment(horizontal="center")
                
                prec = ws.cell(row=row_idx, column=4, value=det.PrecioVenta)
                prec.number_format = '"C$"#,##0.00'
                
                subt = ws.cell(row=row_idx, column=5, value=det.SubTotal)
                subt.number_format = '"C$"#,##0.00'

                ws.row_dimensions[row_idx].outlineLevel = 1
                row_idx += 1
            
            row_idx += 1

    # --- FILA DE TOTALES GENERALES ---
    row_idx += 1 # Espacio
    ws.cell(row=row_idx, column=4, value="TOTALES GENERALES:").font = Font(bold=True)
    
    totales_map = {5: sum_subtotal, 6: sum_propina, 7: sum_desc, 8: sum_total, 10: sum_monto, 11: sum_cambio, 12: sum_2do}
    
    for col, valor in totales_map.items():
        cell = ws.cell(row=row_idx, column=col, value=valor)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=gris_totales)
        cell.number_format = '"C$"#,##0.00'
        cell.border = Border(top=Side(style='medium'))
        
    # --- NUEVA SECCIÓN: TABLA EXTRA DE PROPINAS POR USUARIO ---
    row_idx += 3  # Espacio después de los totales generales
    col_inicio = 12 # Columna L
    col_fin = 13    # Columna M
    
    # Título del Resumen Combinado (Desplazado 11 celdas a la derecha)
    ws.merge_cells(start_row=row_idx, start_column=col_inicio, end_row=row_idx, end_column=col_fin)
    celda_titulo_resumen = ws.cell(row=row_idx, column=col_inicio, value="RESUMEN DE PROPINAS POR USUARIO")
    celda_titulo_resumen.style = "header_style"
    ws.row_dimensions[row_idx].height = 20 
    
    row_idx += 1
    
    # Encabezados de la mini tabla (L y M)
    headers_resumen = ["Usuario/Cargo", "Total de propinas"]
    for i, h in enumerate(headers_resumen):
        c = ws.cell(row=row_idx, column=col_inicio + i, value=h)
        c.style = "header_style"
        c.font = Font(bold=True, size=10, color=blanco)

    row_idx += 1
    
    # Datos del resumen
    for usuario, total_p in propinas_por_usuario.items():
        # Celda de Nombre de Usuario (Columna L)
        c_user = ws.cell(row=row_idx, column=col_inicio, value=usuario)
        c_user.border = borde_detalle
        c_user.alignment = Alignment(horizontal="left")
        
        # Celda de Monto (Columna M)
        c_monto = ws.cell(row=row_idx, column=col_fin, value=total_p)
        c_monto.number_format = '"C$"#,##0.00'
        c_monto.border = borde_detalle
        c_monto.alignment = Alignment(horizontal="right")
        
        row_idx += 1

    # --- AJUSTE AUTOMÁTICO DE COLUMNAS (Tu método solicitado) ---
    for col_idx in range(1, len(columnas) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        # Revisamos todas las celdas de la columna para encontrar el valor más largo
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    # Si es fecha o moneda, le damos un margen extra fijo
                    longitud = len(str(cell.value))
                    if longitud > max_length:
                        max_length = longitud
        
        ws.column_dimensions[column_letter].width = min(max_length + 5, 60)
        
    ws.freeze_panes = "B3"

    return wb

def exportar_pdf_ordenes(ordenes, request, incluir_detalles=False, nombre_archivo="ORDENES"):
    buffer = io.BytesIO()
    pagesize = landscape(letter)
    width, height = pagesize

    doc = SimpleDocTemplate(
        buffer,
        pagesize=pagesize,
        rightMargin=40,
        leftMargin=40,
        topMargin=100,
        bottomMargin=60
    )

    elementos = []
    styles = getSampleStyleSheet()
    
    header_label_style = ParagraphStyle(name="LabelStyle", fontSize=10, fontName="Helvetica-Bold", textColor=colors.HexColor("#8e0000"))
    cell_style = ParagraphStyle(name="CellStyle", fontSize=9, leading=11)
    total_style = ParagraphStyle(name="TotalStyle", fontSize=10, fontName="Helvetica-Bold", textColor=colors.black)
    
    # # # # # RESUMEN # # # # #
    
    # --- NUEVA SECCIÓN: CÁLCULOS PREVIOS ---
    sum_subtotal = sum_propina = sum_desc = sum_total_pagar = sum_monto = sum_cambio = sum_2do = Decimal(0)
    propinas_por_usuario = {}

    for o in ordenes:
        # Acumulación de totales generales
        sum_subtotal += Decimal(str(o.Total or 0))
        sum_propina += Decimal(str(o.Propina or 0))
        sum_desc += Decimal(str(o.Descuento or 0))
        sum_total_pagar += Decimal(str(o.TotalPagar or 0))
        sum_monto += Decimal(str(o.Monto or 0))
        sum_cambio += Decimal(str(o.Cambio or 0))
        sum_2do += Decimal(str(o.SegundoMonto or 0))

        # Agrupación por usuario
        creador_nombre = f"{o.IdUsuario.Nombres} ({o.IdUsuario.IdCargo.Nombre})"
        propinas_por_usuario[creador_nombre] = propinas_por_usuario.get(creador_nombre, Decimal(0)) + Decimal(str(o.Propina or 0))

    # --- TABLA 1: RESUMEN GENERAL DE TOTALES ---
    elementos.append(Paragraph("RESUMEN GENERAL DE VENTAS", header_label_style))
    elementos.append(Spacer(1, 8))
    
    # El ancho total disponible en horizontal (Landscape Letter) es ~732 puntos con márgenes de 40.
    ancho_total = 732 
    ancho_col_resumen = ancho_total / 4 # Dividido en 4 columnas iguales

    resumen_data = [
        [Paragraph("Subtotal Total", header_label_style), Paragraph("Total Propinas", header_label_style), 
         Paragraph("Total Descuentos", header_label_style), Paragraph("TOTAL PAGADO", header_label_style)],
        [f"C$ {sum_subtotal:,.2f}", f"C$ {sum_propina:,.2f}", f"C$ {sum_desc:,.2f}", Paragraph(f"C$ {sum_total_pagar:,.2f}", total_style)],
        [Paragraph("Total Montos", header_label_style), Paragraph("Total Cambios", header_label_style), 
         Paragraph("Total 2do Monto", header_label_style), ""] ,
        [f"C$ {sum_monto:,.2f}", f"C$ {sum_cambio:,.2f}", f"C$ {sum_2do:,.2f}", ""]
    ]
    
    tabla_resumen = Table(resumen_data, colWidths=[ancho_col_resumen] * 4)
    tabla_resumen.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
        ('BACKGROUND', (0, 2), (-1, 2), colors.whitesmoke),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 1), (-1, 1), 'LEFT'), # Alineación de datos
        ('ALIGN', (0, 3), (-1, 3), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elementos.append(tabla_resumen)
    elementos.append(Spacer(1, 20))

    # --- TABLA 2: RESUMEN DE PROPINAS POR USUARIO (Ocupando 100%) ---
    elementos.append(Paragraph("RESUMEN DE PROPINAS POR USUARIO", header_label_style))
    elementos.append(Spacer(1, 8))
    
    propinas_data = [[Paragraph("USUARIO / CARGO", header_label_style), Paragraph("TOTAL RECAUDADO EN PROPINAS", header_label_style)]]
    for usuario, total_p in propinas_por_usuario.items():
        propinas_data.append([usuario.upper(), f"C$ {total_p:,.2f}"])

    # Aquí le damos el 70% del ancho a la descripción y el 30% al monto para que sumen el 100%
    tabla_propinas = Table(propinas_data, colWidths=[ancho_total * 0.5, ancho_total * 0.5])
    tabla_propinas.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f5f0f0")),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'), # Montos a la derecha
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
    ]))
    elementos.append(tabla_propinas)
    elementos.append(Spacer(1, 25))
    elementos.append(HRFlowable(width="100%", thickness=2, color="#8e0000", spaceAfter=20))
    
    # # # # # # # # # # # # # #

    titulo_texto = "REPORTE DETALLADO DE ÓRDENES" if incluir_detalles else "REPORTE GENERAL DE ÓRDENES"
    
    # 1. Definir el nuevo estilo (puedes poner esto junto a tus otros styles)
    ordenes_title_style = ParagraphStyle(
        name="OrdenesTitleStyle", 
        fontSize=14,               # Más grande que el anterior
        fontName="Helvetica-Bold", 
        textColor=colors.HexColor("#8e0000"),
        alignment=1,               # 0=Izquierda, 1=Centro, 2=Derecha
        spaceBefore=20,            # Espacio arriba del título
        spaceAfter=15              # Espacio debajo del título
    )

    # 2. Agregar el elemento justo antes del bucle 'for orden in ordenes:'
    elementos.append(Paragraph("ÓRDENES", ordenes_title_style))
    
    elementos.append(Spacer(1, 3))
    elementos.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#8e0000"), spaceBefore=5, spaceAfter=15))
    
    for orden in ordenes:
        # 1. Preparación de datos del Creador
        u = orden.IdUsuario
        nombre_responsable = f"{u.Nombres} {u.Apellidos}".strip() or u.username
        cargo_responsable = u.IdCargo.Nombre if u.IdCargo else "No asignado"
        
        # 2. Bloque de Identificación y Totales (Se mantiene similar pero ajustamos anchos)
        color_estado = "green" if orden.Estado == "0" else "black"
        estado_html = f'<font color="{color_estado}">●</font> {orden.get_Estado_display()}'
        fecha_fmt = timezone.localtime(orden.Fecha).strftime("%d/%m/%Y %H:%M")
        mesas = " - ".join(f"#{m.IdMesa.Numero}" for m in orden.Mesas.all())

        # Definimos los anchos de columna una sola vez para asegurar consistencia
        col_widths = [110, 180, 150, 290] # Suma total: 730

        data_orden = [
            # Bloque 1: Identificación
            [Paragraph("N° Orden", header_label_style), Paragraph("Fecha / Hora", header_label_style), Paragraph("Estado", header_label_style), Paragraph("Área / Mesas", header_label_style)],
            [orden.Id, fecha_fmt, Paragraph(estado_html, cell_style), f"{orden.IdAreaDeMesa.Nombre if orden.IdAreaDeMesa else 'N/A'} ({mesas})"],
            
            # Bloque 2: Totales
            [Paragraph("Subtotal", header_label_style), Paragraph("Propina", header_label_style), Paragraph("Descuento", header_label_style), Paragraph("TOTAL A PAGAR", header_label_style)],
            [f"C$ {orden.Total}", f"C$ {orden.Propina}", f"C$ {orden.Descuento}", Paragraph(f"C$ {orden.TotalPagar}", total_style)],
            
            # Bloque 3: Detalles del Pago
            [Paragraph("Método de Pago", header_label_style), Paragraph("Monto Recibido", header_label_style), Paragraph("Cambio", header_label_style), Paragraph("2do Monto (Pago Mixto)", header_label_style)],
            [orden.get_MetodoPago_display(), f"C$ {orden.Monto}", f"C$ {orden.Cambio}", f"C$ {orden.SegundoMonto}" if orden.MetodoPago == "4" else "N/A"],

            # Bloque 4: Creador (Alineado con las celdas superiores)
            # Usamos "" para las celdas que serán "absorbidas" por el SPAN
            [Paragraph("ATENDIDO POR:", header_label_style), "", Paragraph("CARGO DEL RESPONSABLE:", header_label_style), ""],
            [Paragraph(nombre_responsable.upper(), cell_style), "", Paragraph(cargo_responsable.upper(), cell_style), ""]
        ]

        tabla_cabecera = Table(data_orden, colWidths=col_widths)
        tabla_cabecera.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            # Fondos de etiquetas
            ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
            ('BACKGROUND', (0, 2), (-1, 2), colors.whitesmoke),
            ('BACKGROUND', (0, 4), (-1, 4), colors.whitesmoke),
            ('BACKGROUND', (0, 6), (-1, 6), colors.whitesmoke), # Fondo para etiquetas de creador
            
            # --- COMBINACIÓN DE CELDAS (SPAN) ---
            # Combinamos las dos primeras y las dos últimas de las filas 6 y 7
            ('SPAN', (0, 6), (1, 6)), # Atendido por (Etiqueta)
            ('SPAN', (2, 6), (3, 6)), # Cargo (Etiqueta)
            ('SPAN', (0, 7), (1, 7)), # Nombre (Valor)
            ('SPAN', (2, 7), (3, 7)), # Cargo (Valor)
            
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        elementos.append(tabla_cabecera)

        # TABLA 2: DETALLES DE PRODUCTOS
        if incluir_detalles:
            elementos.append(Spacer(1, 5))
            detalles_data = [["Nombre del consumible", "Precio Unitario", "Cantidad",  "Subtotal"]]
            detalles_activos = orden.Detalles.filter(EsActivo="1")

            for det in detalles_activos:
                detalles_data.append([
                    Paragraph(det.IdPlatillo.Nombre, cell_style),
                    f"C$ {det.PrecioVenta}",
                    det.Cantidad,
                    f"C$ {det.SubTotal}"
                ])

            if not detalles_activos:
                detalles_data.append(["-", "Sin detalles registrados", "-", "-"])

            tabla_detalle = Table(detalles_data, colWidths=[420, 120, 70, 120])
            tabla_detalle.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f5f0f0")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#a05047")),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (2, 0), (2, -1), 'CENTER'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
            ]))
            elementos.append(tabla_detalle)

        # --- SEPARADOR ENTRE ÓRDENES ---
        elementos.append(Spacer(1, 15))
        elementos.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#8e0000"), spaceBefore=5, spaceAfter=15))
        # -------------------------------

    # --- ENCABEZADO Y PIE ---
    # (Se mantiene igual a tu última versión...)
    def encabezado_pie(canvas, doc):
        canvas.saveState()
        logo_path = os.path.join("static", "img", "NuevoLogoFondoBlanco.jpg")
        if os.path.exists(logo_path):
            canvas.drawImage(logo_path, doc.leftMargin, height - 80, width=60, height=60, preserveAspectRatio=True)

        canvas.setFont("Helvetica-Bold", 18)
        canvas.setFillColor(colors.HexColor("#8e0000"))
        canvas.drawCentredString(width / 2, height - 62, titulo_texto.upper())
        
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(colors.grey)
        fecha_emision = timezone.localtime().strftime("%d/%m/%Y %H:%M")
        canvas.drawRightString(width - doc.rightMargin, height - 50, f"Emisión: {fecha_emision}")
        canvas.drawRightString(width - doc.rightMargin, height - 65, f"Usuario: {request.user.username}")

        canvas.setStrokeColor(colors.HexColor("#8e0000"))
        canvas.setLineWidth(2)
        canvas.line(doc.leftMargin, height - 90, width - doc.rightMargin, height - 90)

        canvas.setFont("Helvetica", 8)
        canvas.drawString(doc.leftMargin, 35, "Documento generado por el sistema.")
        canvas.drawRightString(width - doc.rightMargin, 35, f"Página {doc.page}")
        canvas.restoreState()

    doc.build(elementos, onFirstPage=encabezado_pie, onLaterPages=encabezado_pie)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}.pdf"'
    return response

#endregion Ordenes

#region Platillos

def ExportarPlatillo(request):
    if not request.user.is_authenticated:
        return render(request, "login.html")
    
    if request.user.IdCargo.Nombre == "Armador" or request.user.IdCargo.Nombre == "Mesero":
        return redirect("/")
    
    try:
        tipoExportacion = request.GET.get("Tipo")
        
        if tipoExportacion not in ("1","2"):
            return JsonResponse({
                "status": "error",
                "message": "Tipo de exportación inválido"
            }, status=400)
        
        if tipoExportacion == "1":
            return exportar_excel_platillos()
        
        if tipoExportacion == "2":
            return exportar_pdf_platillo(request)
        
        return JsonResponse({
            "status": "error",
            "message": "Exportación no implementada"
        }, status=400)
    except Exception:
        print("\n\n#################### E X C E P C I O N ########################")
        print("-------------------- 'exportar platillo' --------------------")
        print(traceback.format_exc())
        print("#############################################################\n\n")
        return JsonResponse({
            "status": "error",
            "message": "Error interno al exportar consumibles"
        }, status=500)
        
def exportar_excel_platillos():
    columnas = ['Nombre consumible', 'Precio', 'Tipo de consumible', 'Descripcion', 'Estado']
    datos = []

    for nombre, precio, tipo, desc, es_activo in Platillo.objects.values_list(
        'Nombre', 'Precio', 'IdTipoPlatillo__Nombre', 'Descripcion', 'EsActivo'
    ):
        estado_symbol = '✅ Activo' if es_activo in (1, '1', True) else '⛔ Inactivo'
        datos.append([
            nombre,
            precio,
            tipo,
            desc,
            estado_symbol
        ])

    wb = exportar_excel_datos("CONSUMIBLES", columnas, datos, "Consumibles")

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = 'CONSUMIBLES_' + timezone.localtime().strftime('%d-%m-%Y') + '.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    
    return response

def exportar_pdf_platillo(request):
    columnas = ['Nombre consumible', 'Precio', 'Tipo de consumible', 'Descripcion', 'Estado']
    filas = []

    for nombre, precio, tipo, desc, es_activo in Platillo.objects.values_list('Nombre', 'Precio', 'IdTipoPlatillo__Nombre', 'Descripcion', 'EsActivo'):
        estado_symbol = '✓ Activo' if es_activo in (1, '1', True) else '✗ Inactivo'
        filas.append([
            nombre,
            precio,
            tipo,
            desc,
            estado_symbol
        ])

    return generar_pdf_tabla(
        titulo="CONSUMIBLES",
        columnas=columnas,
        filas=filas,
        nombre_archivo="CONSUMIBLES_" + timezone.localtime().strftime('%d-%m-%Y'),
        ancho_columnas=[100, 50, 100, 200, 70],
        wrap_columns=[0, 2, 3],
        usuario=request.user.username
    )
        
#endregion Platillos

#region TipoPlatillos

def ExportarTipoPlatillo(request):
    if not request.user.is_authenticated:
        return render(request, "login.html")
    
    if request.user.IdCargo.Nombre == "Armador" or request.user.IdCargo.Nombre == "Mesero":
        return redirect("/")

    try:
        tipoExportacion = request.GET.get("Tipo")
        
        if tipoExportacion not in ("1","2"):
            return JsonResponse({
                "status": "error",
                "message": "Tipo de exportación inválido"
            }, status=400)
        
        if tipoExportacion == "1":
            return exportar_excel_tipo_platillo()
        
        if tipoExportacion == "2":
            return exportar_pdf_tipo_platillo(request)
        
        return JsonResponse({
            "status": "error",
            "message": "Exportación no implementada"
        }, status=400)
    except Exception:
        print("\n\n#################### E X C E P C I O N ########################")
        print("---------------- 'exportar tipo platillo' ----------------")
        print(traceback.format_exc())
        print("#############################################################\n\n")
        
        return JsonResponse({
            "status": "error",
            "message": "Error interno al exportar consumibles"
        }, status=500)
        
def exportar_excel_tipo_platillo():
    columnas = ['Nombre', 'Estado']
    datos = []

    for nombre, es_activo in TipoPlatillo.objects.values_list('Nombre', 'EsActivo'):
        estado_symbol = '✅ Activo' if es_activo in (1, '1', True) else '⛔ Inactivo'
        datos.append([
            nombre,
            estado_symbol
        ])

    wb = exportar_excel_datos("TIPO DE CONSUMIBLE", columnas, datos, "Tipos de consumible")

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = 'TIPO_CONSUMIBLE_' + timezone.localtime().strftime('%d-%m-%Y') + '.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def exportar_pdf_tipo_platillo(request):
    columnas = ["Nombre", "Estado"]
    filas = []

    for nombre, es_activo in TipoPlatillo.objects.values_list("Nombre", "EsActivo"):
        estado = "✓ Activo" if es_activo == "1" else "✗ Inactivo"
        filas.append([nombre, estado])

    return generar_pdf_tabla(
        titulo="TIPOS DE CONSUMIBLE",
        columnas=columnas,
        filas=filas,
        nombre_archivo="TIPO_CONSUMIBLE_" + timezone.localtime().strftime('%d-%m-%Y'),
        ancho_columnas=[300, 150],
        usuario=request.user.username
    )
        
#endregion TipoPlatillos

#region Personal
        
def ExportarPersonal(request):
    if not request.user.is_authenticated:
        return render(request, "login.html")
    
    if request.user.IdCargo.Nombre == "Armador" or request.user.IdCargo.Nombre == "Mesero":
        return redirect("/")

    try:
        tipoExportacion = request.GET.get("Tipo")
        
        if tipoExportacion not in ("1","2"):
            return JsonResponse({
                "status": "error",
                "message": "Tipo de exportación inválido"
            }, status=400)
        
        # Para exportar en excel
        if tipoExportacion == "1":
            return exportar_excel_personal()
        
        # Para exportar en pdf
        if tipoExportacion == "2":
            return exportar_pdf_personal(request)
        
        return JsonResponse({
            "status": "error",
            "message": "Exportación no implementada"
        }, status=400)

    except Exception:
        print("\n\n#################### E X C E P C I O N ########################")
        print("---------------- 'exportar tipo platillo' ----------------")
        print(traceback.format_exc())
        print("#############################################################\n\n")
        
def exportar_excel_personal():
    columnas = ['Id', 'Nombres y apellidos', 'Usuario', 'Cargo', 'Telefono', 'Correo', 'Estado']
    datos = []

    for id, nombre, apellido, nombre_user, nombre_cargo, telefono, correo, es_activo in Usuario.objects.values_list('Id', 'Nombres', 'Apellidos', 'username', 'IdCargo__Nombre', 'Telefono', 'email', 'EsActivo'):
        estadoData = '✅ Activo' if es_activo in (1, '1', True) else '⛔ Dado de baja'
        datos.append([
            id,
            nombre + " " + apellido,
            nombre_user,
            nombre_cargo,
            telefono,
            correo,
            estadoData
        ])

    wb = exportar_excel_datos("PERSONAL", columnas, datos, "Personal")

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = 'PERSONAL_' + timezone.localtime().strftime('%d-%m-%Y') + '.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def exportar_pdf_personal(request):
    columnas = ['Id', 'Nombres y apellidos', 'Usuario', 'Cargo', 'Telefono', 'Correo', 'Estado']
    filas = []

    for id, nombre, apellido, nombre_user, nombre_cargo, telefono, correo, es_activo in Usuario.objects.values_list('Id', 'Nombres', 'Apellidos', 'username', 'IdCargo__Nombre', 'Telefono', 'email', 'EsActivo'):
        estadoData = '✓ Activo' if es_activo in (1, '1', True) else '✗ Dado de baja'
        filas.append([
            id,
            nombre + " " + apellido,
            nombre_user,
            nombre_cargo,
            telefono,
            correo,
            estadoData
        ])

    return generar_pdf_tabla(
        titulo="PERSONAL",
        columnas=columnas,
        filas=filas,
        nombre_archivo="PERSONAL_" + timezone.localtime().strftime('%d-%m-%Y'),
        ancho_columnas=[50, 150, 100, 100, 60, 170, 80],
        usuario=request.user.username,
        wrap_columns=[1],
        horizontal=True
    )
        
#endregion Personal

#region Caja

def ExportarArqueo(request):
    if not request.user.is_authenticated:
        return render(request, "login.html")
    
    if request.user.IdCargo.Nombre == "Armador" or request.user.IdCargo.Nombre == "Mesero":
        return redirect("/")
    
    try:
        formato = request.GET.get('Tipo', '1')  # 1: Excel, 2: PDF
        dias = request.GET.get('Dias', '1')  # 1: Excel, 2: PDF
        
        if formato not in ("1","2"):
            return JsonResponse({
                "status": "error",
                "message": "Tipo de exportación inválido"
            }, status=400)
        
        hoy = timezone.localdate()
        arqueo = Arqueo.objects.filter(Fecha=hoy).first()
        
        inicio_dia = timezone.make_aware(datetime.combine(hoy, time.min))
        fin_dia = timezone.make_aware(datetime.combine(hoy, time.max))

        # 4. Consultas de Órdenes usando el rango (range)
        # Esto obliga a Django a convertir los UTC de la BD a Nicaragua antes de comparar
        ordenes = Orden.objects.filter(
            UltimaModificacion__range=(inicio_dia, fin_dia), 
            Estado="0", 
            EsActivo="1"
        )
        
        # Mixto
        efectivo_mixto = ordenes.filter(MetodoPago="4").aggregate(total=Sum('Monto') - Sum('Cambio'))['total'] or 0
        tarjeta_mixta = ordenes.filter(MetodoPago="4").aggregate(total=Sum('SegundoMonto'))['total'] or 0
        cantidad_mixto = ordenes.filter(MetodoPago="4").count()
        
        # Efectivo
        efectivo_puro = ordenes.filter(MetodoPago="1").aggregate(total=Sum('TotalPagar'))['total'] or 0
        cantidad_efectivo = ordenes.filter(MetodoPago="1").count()
        
        # Efectivo + Mixto
        total_efectivo = efectivo_puro + efectivo_mixto
        cantidad_efectivo_mixto = cantidad_efectivo + cantidad_mixto
        
        # Tarjeta
        tarjeta_pura = ordenes.filter(MetodoPago="2").aggregate(total=Sum('TotalPagar'))['total'] or 0
        cantidad_tarjeta = ordenes.filter(MetodoPago="2").count()
        
        # Tarjeta + Mixto
        total_tarjeta = tarjeta_pura + tarjeta_mixta
        cantidad_tarjeta_mixto = cantidad_tarjeta + cantidad_mixto
        
        # Transferencia
        total_transferencia = ordenes.filter(MetodoPago="3").aggregate(total=Sum('TotalPagar'))['total'] or 0
        cantidad_transferencia = ordenes.filter(MetodoPago="3").count()
        
        metodos = {
            '1': {'nombre': 'Efectivo', 'total': Decimal(total_efectivo), 'cantidad': cantidad_efectivo_mixto},
            '2': {'nombre': 'Tarjeta', 'total': Decimal(total_tarjeta), 'cantidad': cantidad_tarjeta_mixto},
            '3': {'nombre': 'Transferencia', 'total': Decimal(total_transferencia), 'cantidad': cantidad_transferencia},
        }

        total_general_monto = Decimal(total_efectivo) + Decimal(total_tarjeta) + Decimal(total_transferencia)
        total_general_cantidad = cantidad_efectivo + cantidad_tarjeta + cantidad_mixto + cantidad_transferencia

        # Preparar filas para las tablas
        filas_metodos = []
        for m in metodos.values():
            filas_metodos.append([m['nombre'], m['cantidad'], m['total']])
        
        # Fila de totales
        filas_metodos.append(["TOTALES", total_general_cantidad, total_general_monto])

        # Datos informativos del Arqueo
        datos_info = [
            ["Fecha Arqueo", arqueo.Fecha.strftime("%d/%m/%Y %H:%M")],
            ["Usuario Apertura", arqueo.IdUsuarioApertura.username],
            ["Usuario Cierre", arqueo.IdUsuarioCierre.username if arqueo.IdUsuarioCierre else "N/A"],
            ["Monto Inicial", arqueo.MontoInicial],
            ["Monto Teórico", arqueo.MontoFinalTeorico],
            ["Monto Real", arqueo.MontoFinalReal],
            ["Diferencia", arqueo.Diferencia],
        ]

        titulo = f"REPORTE DE ARQUEO DE CAJA #{arqueo.Id}"
        nombre_archivo = f"Arqueo_{arqueo.Id}_{arqueo.Fecha.strftime('%Y%m%d')}"

        if formato == '1':
            # --- EXCEL ---
            wb = Workbook()
            
            if "header_style" not in wb.style_names:
                header_style = NamedStyle(name="header_style")
                wb.add_named_style(header_style)
            
            ws = wb.active
            ws.title = "Detalle Arqueo"
            
            # Usamos tu función 'poblar_hoja_existente' o similar
            # Primero info general
            poblar_hoja_existente(ws, titulo, ["Descripción", "Valor"], datos_info)
            
            # Agregamos la tabla de métodos de pago más abajo
            row_start = len(datos_info) + 5
            ws.cell(row=row_start-1, column=1, value="RESUMEN POR MÉTODO DE PAGO").font = Font(bold=True)
            
            columnas_m = ["Método de Pago", "Cant. Órdenes", "Total"]
            # Aquí podrías adaptar una función para insertar tablas en posiciones específicas
            for i, h in enumerate(columnas_m, 1):
                cell = ws.cell(row=row_start, column=i, value=h)
                cell.style = "header_style" # Usando el estilo que definiste antes

            for idx, fila in enumerate(filas_metodos):
                for col_idx, valor in enumerate(fila, 1):
                    c = ws.cell(row=row_start + 1 + idx, column=col_idx, value=valor)
                    if col_idx == 3: c.number_format = '"C$"#,##0.00'

            return descargar_excel(wb, nombre_archivo)

        else:
            # --- PDF ---
            # Combinamos toda la información en una estructura de filas para generar_pdf_tabla
            # O mejor, usamos una lógica personalizada similar a tu exportar_pdf_ordenes
            columnas_pdf = ["CATEGORÍA / MÉTODO", "CANTIDAD", "MONTO"]
            filas_pdf = []
            
            # Mezclamos la info del arqueo con los totales para una sola tabla clara
            for d in datos_info:
                filas_pdf.append([d[0], "-", f"C$ {d[1]}" if isinstance(d[1], (int, float, Decimal)) else d[1]])
            
            filas_pdf.append(["", "", ""]) # Separador
            filas_pdf.append(["RESUMEN DE PAGOS", "", ""])
            
            for m in filas_metodos:
                filas_pdf.append([m[0], m[1], f"C$ {m[2]:,.2f}" if isinstance(m[2], Decimal) else m[2]])

            return generar_pdf_tabla(
                titulo=titulo,
                columnas=columnas_pdf,
                filas=filas_pdf,
                nombre_archivo=nombre_archivo,
                usuario=request.user.username,
                ancho_columnas=[250, 100, 150]
            )
    except:
        print("\n\n#################### E X C E P C I O N ########################")
        print("---------------- 'exportar arqueo de caja' ----------------")
        print(traceback.format_exc())
        print("#############################################################\n\n")
        
        return JsonResponse({
            "status": "error",
            "message": "Error interno al exportar datos de arqueo"
        }, status=500)

#endregion Caja

#endregion Exportaciones

#region PublicFunctions

def filtrar_ordenes_fechas_areas(fecha_inicio_str, fecha_fin_str, areas_ids, estado = "0"):
    
    listaEstado = estado.split(",")
    
    if not fecha_inicio_str or not fecha_fin_str:
        raise ValidationError("Fechas incompletas")

    # Convertir string → date
    fecha_inicio_date = datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
    fecha_fin_date = datetime.strptime(fecha_fin_str, "%Y-%m-%d").date()

    # Validación lógica
    if fecha_inicio_date > fecha_fin_date:
        raise ValidationError("La fecha inicio no puede ser mayor a la fecha fin")

    # Ajustar horas
    fecha_inicio = timezone.make_aware(datetime.combine(fecha_inicio_date, time.min), timezone.get_current_timezone())   # 00:00:00
    fecha_fin = timezone.make_aware(datetime.combine(fecha_fin_date, time.max), timezone.get_current_timezone())         # 23:59:59.999999
    
    filtros = Q(
        EsActivo="1",
        UltimaModificacion__range=(fecha_inicio, fecha_fin),
        Estado__in=listaEstado
    )

    # Filtrar por áreas de mesa (opcional)
    if areas_ids:
        filtros &= Q(IdAreaDeMesa__in=areas_ids)

    # ------------------------
    # Query final
    # ------------------------
    ordenes = (
        Orden.objects
        .select_related('IdUsuario', 'IdAreaDeMesa')
        .prefetch_related(Prefetch('Detalles'), Prefetch('Mesas', queryset=MesasPorOrden.objects.filter(EsActivo="1").select_related('IdMesa')))
        .filter(filtros)
        .order_by("-Id")
    )
    
    return ordenes

def exportar_excel_datos(titulo, columnas, datos, sheetName):
    # ==== CREAR LIBRO ====
    wb = Workbook()
    ws = wb.active
    ws.title = sheetName

    # ==== ESTILOS ====
    vino_oscuro = "800000"
    blanco = "FFFFFF"
    font_blanca_sans = Font(bold=True, color=blanco, name="Arial")

    borde_resaltado = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    # ==== TÍTULO ====
    ultima_columna = get_column_letter(len(columnas))
    ws.merge_cells(f"A1:{ultima_columna}1")
    celda_titulo = ws["A1"]
    celda_titulo.value = titulo
    celda_titulo.font = font_blanca_sans
    celda_titulo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    celda_titulo.fill = PatternFill("solid", fgColor=vino_oscuro)
    celda_titulo.border = borde_resaltado
    ws.row_dimensions[1].height = 25

    # ==== ENCABEZADOS ====
    for col_idx, header in enumerate(columnas, start=1):
        c = ws.cell(row=2, column=col_idx, value=header)
        c.font = font_blanca_sans
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=vino_oscuro)
        c.border = borde_resaltado
    ws.row_dimensions[2].height = 20

    # ==== DATOS ====
    start_row = 3
    for idx, fila in enumerate(datos, start=0):
        row_index = start_row + idx

        # Color alterno
        fill_color = "FFF5F5" if idx % 2 == 0 else "FFF0E6"

        for col_idx, valor in enumerate(fila, start=1):
            cell = ws.cell(row=row_index, column=col_idx, value=valor)

            # Formato moneda si la columna es "Precio"
            if columnas[col_idx - 1].lower() == "precio" and isinstance(valor, (int, float, Decimal)):
                cell.number_format = u'"C$"#,##0.00'

            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = borde_resaltado

    # ==== AJUSTES VISUALES ====
    ws.freeze_panes = "A3"
    
    for col_idx, header in enumerate(columnas, start=1):
        column_letter = get_column_letter(col_idx)
        max_length = len(str(header))

        for row in ws.iter_rows(
            min_row=2,
            max_row=ws.max_row,
            min_col=col_idx,
            max_col=col_idx
        ):
            cell_value = row[0].value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        ws.column_dimensions[column_letter].width = min(max_length + 4, 50)

    return wb

def poblar_hoja_existente(ws, titulo, columnas, datos):
    vino_oscuro = "800000"
    blanco = "FFFFFF"
    font_blanca_sans = Font(bold=True, color=blanco, name="Arial")

    borde_resaltado = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    ultima_columna = get_column_letter(len(columnas))
    ws.merge_cells(f"A1:{ultima_columna}1")

    ws["A1"].value = titulo
    ws["A1"].font = font_blanca_sans
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor=vino_oscuro)
    ws["A1"].border = borde_resaltado
    ws.row_dimensions[1].height = 25

    for col_idx, header in enumerate(columnas, start=1):
        c = ws.cell(row=2, column=col_idx, value=header)
        c.font = font_blanca_sans
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor=vino_oscuro)
        c.border = borde_resaltado

    start_row = 3
    for idx, fila in enumerate(datos):
        fill_color = "FFF5F5" if idx % 2 == 0 else "FFF0E6"
        for col_idx, valor in enumerate(fila, start=1):
            cell = ws.cell(row=start_row + idx, column=col_idx, value=valor)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = borde_resaltado

    ws.freeze_panes = "A3"
    for i, header in enumerate(columnas, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(len(header) + 5, 15)

def descargar_excel(wb, nombre_archivo):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # Limpiamos el nombre por si trae espacios o caracteres raros
    nombre_limpio = nombre_archivo.replace(" ", "_")
    response["Content-Disposition"] = f'attachment; filename="{nombre_limpio}"'
    wb.save(response)
    return response

## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ##
 ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ##
## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ## ##

def generar_pdf_tabla(*, titulo: str, columnas: list, filas: list, nombre_archivo: str,
    ancho_columnas=None, wrap_columns=None, horizontal=False, usuario=None
):
    if wrap_columns is None: wrap_columns = []
    buffer = io.BytesIO()
    pagesize = landscape(letter) if horizontal else letter
    width, height = pagesize

    # Ajustamos el margen superior para que el contenido no choque con el encabezado elegante
    doc = SimpleDocTemplate(
        buffer,
        pagesize=pagesize,
        rightMargin=40,
        leftMargin=40,
        topMargin=120,  # 👈 Aumentado para dar aire al encabezado
        bottomMargin=60
    )

    elementos = []
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle(name="CellStyle", fontSize=9, leading=12, alignment=0, wordWrap="CJK")

    # --- PROCESAR FILAS ---
    filas_procesadas = [[Paragraph(str(valor) if valor else "", cell_style) if i in wrap_columns else valor 
                         for i, valor in enumerate(fila)] for fila in filas]

    # --- TABLA DE DATOS ---
    # Nota: Eliminamos el título de adentro de la tabla para ponerlo en el encabezado elegante
    data = [columnas] + filas_procesadas
    table = Table(data, colWidths=ancho_columnas, repeatRows=1)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#8e0000")), # Rojo elegante
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elementos.append(table)

    # ======================================================
    # ENCABEZADO TIPO FACTURA (POSICIÓN ABSOLUTA)
    # ======================================================
    def encabezado_pie(canvas, doc):
        canvas.saveState()
        
        # 1. Dibujar el Logo (Posición Absoluta: Top-Izquierda)
        logo_path = os.path.join("static", "img", "NuevoLogoFondoBlanco.jpg")
        if os.path.exists(logo_path):
            # drawImage(ruta, x, y, width, height)
            canvas.drawImage(logo_path, doc.leftMargin, height - 80, width=60, height=60, preserveAspectRatio=True)

        # 2. Título Principal (Centro-Izquierda, al lado del logo)
        canvas.setFont("Helvetica-Bold", 18)
        canvas.setFillColor(colors.HexColor("#a05047"))
        canvas.drawString(doc.leftMargin + 70, height - 62, titulo.upper())
        
        # 3. Bloque de Información (Derecha - Estilo Factura)
        canvas.setFont("Helvetica-Bold", 10)
        canvas.drawRightString(width - doc.rightMargin, height - 40, "REPORTE OFICIAL")
        
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(colors.grey)
        fecha_str = timezone.localtime().strftime("%d/%m/%Y %H:%M")
        canvas.drawRightString(width - doc.rightMargin, height - 55, f"Fecha de emisión: {fecha_str}")
        if usuario:
            canvas.drawRightString(width - doc.rightMargin, height - 67, f"Generado por: {usuario}")

        # 4. Línea decorativa elegante
        canvas.setStrokeColor(colors.HexColor("#8e0000"))
        canvas.setLineWidth(2)
        canvas.line(doc.leftMargin, height - 90, width - doc.rightMargin, height - 90)

        # --- PIE DE PÁGINA ---
        canvas.setFont("Helvetica", 8)
        canvas.setStrokeColor(colors.lightgrey)
        canvas.line(doc.leftMargin, 50, width - doc.rightMargin, 50) # Línea superior al pie
        canvas.drawString(doc.leftMargin, 35, f"Documento generado por el sistema.")
        canvas.drawRightString(width - doc.rightMargin, 35, f"Página {canvas.getPageNumber()}")
        
        canvas.restoreState()

    # --- CONSTRUIR ---
    doc.build(elementos, onFirstPage=encabezado_pie, onLaterPages=encabezado_pie)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}.pdf"'
    return response

#endregion PublicFunctions